"""
utils/excel_exporter.py
-----------------------
Builds the 12-sheet KKC JE Testing Tool Excel workpaper from the
standardised module result dicts returned by each analysis module.

Sheet order (per CLAUDE.md):
  1  Summary Dashboard
  2  Benfords Analysis
  3  Duplicates
  4  Round Numbers
  5  After Hours Weekend
  6  Period End Entries
  7  Missing Narration
  8  Unusual Acct Combos
  9  User Analysis
  10 Keyword Flags
  11 Full Population
  12 Parameters

Public API:
  build_workbook(full_df, module_results, params, output_path) -> str
  generate_filename(client_name, period, run_date) -> str
"""

from __future__ import annotations

import os
from datetime import datetime
from typing import Any

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows

try:
    from config import (
        KKC_GREEN,
        KKC_LIGHT_GREY,
        RISK_RED,
        RISK_ORANGE,
        OUTPUT_PREFIX,
        DISPLAY_DATE_FORMAT,
    )
except ImportError:
    KKC_GREEN           = "#7CB542"
    KKC_LIGHT_GREY      = "#F2F2F2"
    RISK_RED            = "#FF0000"
    RISK_ORANGE         = "#FFA500"
    OUTPUT_PREFIX       = "KKC_JE_"
    DISPLAY_DATE_FORMAT = "%d %B %Y"


# ---------------------------------------------------------------------------
# Module → sheet name mapping (fixed order for workbook)
# ---------------------------------------------------------------------------

# List of (module_name_key, excel_sheet_name) in the required output order.
# module_name_key must match the "module_name" value each module returns.
_MODULE_SHEET_MAP: list[tuple[str, str]] = [
    ("Benford's Law",                  "Benfords Analysis"),
    ("Duplicate Detection",            "Duplicates"),
    ("Round Number Detection",         "Round Numbers"),
    ("After-Hours & Weekend Postings", "After Hours Weekend"),
    ("Period-End Entry Concentration", "Period End Entries"),
    ("Missing Narration",              "Missing Narration"),
    ("Unusual Account Combinations",   "Unusual Acct Combos"),
    ("User Analysis",                  "User Analysis"),
    ("Keyword Flags",                  "Keyword Flags"),
]


# ---------------------------------------------------------------------------
# Colour / style helpers
# ---------------------------------------------------------------------------

def _hex(colour: str) -> str:
    """Strip leading '#' from a hex colour string (openpyxl needs bare hex)."""
    return colour.lstrip("#")


# Pre-built fill objects
_GREEN_FILL  = PatternFill("solid", fgColor=_hex(KKC_GREEN))
_GREY_FILL   = PatternFill("solid", fgColor=_hex(KKC_LIGHT_GREY))
_WHITE_FILL  = PatternFill("solid", fgColor="FFFFFF")
_RED_FILL    = PatternFill("solid", fgColor=_hex(RISK_RED))
_ORANGE_FILL = PatternFill("solid", fgColor=_hex(RISK_ORANGE))

# Pre-built font objects
_HEADER_FONT = Font(name="Arial", size=12, bold=True, color="FFFFFF")
_TITLE_FONT  = Font(name="Arial", size=14, bold=True)
_BOLD_FONT   = Font(name="Arial", size=11, bold=True)
_NORMAL_FONT = Font(name="Arial", size=11)
_SMALL_FONT  = Font(name="Arial", size=10)

# Alignments
_CENTRE = Alignment(horizontal="center", vertical="center")
_LEFT   = Alignment(horizontal="left",   vertical="center")
_RIGHT  = Alignment(horizontal="right",  vertical="center")

# Number formats
_DATE_FMT   = "DD-MMM-YYYY"
_AMOUNT_FMT = '₹ #,##,##0.00'     # Indian number format with ₹ prefix


def _risk_fill(risk: str) -> PatternFill:
    """Return the fill object for a given risk rating."""
    return {"High": _RED_FILL, "Medium": _ORANGE_FILL, "Low": _GREEN_FILL}.get(
        risk, _WHITE_FILL
    )


def _risk_font_white() -> Font:
    """Bold white font for use on coloured risk backgrounds."""
    return Font(name="Arial", size=11, bold=True, color="FFFFFF")


def _tab_colour(risk: str) -> str:
    """Return a bare-hex tab colour for a sheet based on its risk rating."""
    return {
        "High":   _hex(RISK_RED),
        "Medium": _hex(RISK_ORANGE),
        "Low":    _hex(KKC_GREEN),
    }.get(risk, "FFFFFF")


# ---------------------------------------------------------------------------
# Low-level worksheet helpers
# ---------------------------------------------------------------------------

def _freeze_top(ws) -> None:
    """Freeze the top row of a worksheet."""
    ws.freeze_panes = "A2"


def _autofit(ws, max_width: int = 50) -> None:
    """
    Auto-fit all column widths in a worksheet, capped at max_width.

    Args:
        ws:        openpyxl Worksheet.
        max_width: Maximum column width in characters.
    """
    for col in ws.columns:
        col_letter = col[0].column_letter
        max_len = 0
        for cell in col:
            try:
                if cell.value is not None:
                    max_len = max(max_len, len(str(cell.value)))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max_len + 4, max_width)


def _apply_header_row(ws, row_num: int, n_cols: int) -> None:
    """Apply KKC Green fill and white bold text to a single header row."""
    for c in range(1, n_cols + 1):
        cell = ws.cell(row=row_num, column=c)
        cell.fill      = _GREEN_FILL
        cell.font      = _HEADER_FONT
        cell.alignment = _CENTRE


def _apply_alternating(ws, start_row: int, end_row: int, n_cols: int) -> None:
    """Apply alternating white / light-grey fill to data rows."""
    for r in range(start_row, end_row + 1):
        fill = _GREY_FILL if (r - start_row) % 2 == 1 else _WHITE_FILL
        for c in range(1, n_cols + 1):
            ws.cell(row=r, column=c).fill = fill


def _is_date_col(col_name: str, dtype) -> bool:
    """True if column name contains 'date' or dtype is datetime."""
    return "date" in str(col_name).lower() or "datetime" in str(dtype).lower()


def _is_amount_col(col_name: str) -> bool:
    """True if column name looks like a monetary amount column."""
    lower = str(col_name).lower()
    return any(kw in lower for kw in ("amount", "total_amount", "avg_amount"))


def _safe_value(v: Any) -> Any:
    """
    Convert pandas NA / NaT to None so openpyxl does not error.

    Also converts pandas Period objects to strings, and datetime
    objects are left as-is (openpyxl handles them natively).
    """
    if v is None:
        return None
    try:
        if isinstance(v, pd.Period):
            return str(v)
        if pd.isna(v):
            return None
    except (TypeError, ValueError):
        pass
    return v


def _section_header(ws, row_num: int, text: str, n_cols: int = 6) -> None:
    """
    Write a single-cell section header with green fill and white bold text.

    Args:
        ws:      Worksheet.
        row_num: Row to write the header on.
        text:    Header text.
        n_cols:  Number of columns to fill with the green colour.
    """
    for c in range(1, n_cols + 1):
        cell = ws.cell(row=row_num, column=c)
        cell.fill      = _GREEN_FILL
        cell.font      = _HEADER_FONT
        cell.alignment = _LEFT
    ws.cell(row=row_num, column=1).value = text


def _write_kv_block(ws, pairs: list[tuple[str, Any]], start_row: int) -> int:
    """
    Write a vertical key/value block (label in col A, value in col B).

    Args:
        ws:        Worksheet.
        pairs:     List of (label, value) tuples.
        start_row: First row to write.

    Returns:
        Row number immediately after the last written row.
    """
    for i, (label, value) in enumerate(pairs):
        r = start_row + i
        lc = ws.cell(row=r, column=1, value=label)
        vc = ws.cell(row=r, column=2, value=value)
        lc.font      = _BOLD_FONT
        vc.font      = _NORMAL_FONT
        lc.alignment = _LEFT
        vc.alignment = _LEFT
    return start_row + len(pairs)


# ---------------------------------------------------------------------------
# DataFrame writer
# ---------------------------------------------------------------------------

def _write_df(
    ws,
    df: pd.DataFrame,
    start_row: int,
    risk_col_name: str | None = None,
) -> int:
    """
    Write a DataFrame to a worksheet with full KKC formatting.

    Applies:
      - KKC Green header row
      - Alternating white / light-grey data rows
      - DD-MMM-YYYY format for date columns
      - Indian ₹ format for amount columns
      - Risk-colour fill for a nominated risk column

    Args:
        ws:            Worksheet.
        df:            DataFrame to write (must not be empty for formatting
                       to apply, though an empty-state message is written if
                       it is empty).
        start_row:     First row to write the header on.
        risk_col_name: Optional column name to apply risk colour coding to.

    Returns:
        Row number immediately after the last written row.
    """
    if df.empty:
        ws.cell(row=start_row, column=1, value="(No flagged entries for this module)")
        ws.cell(row=start_row, column=1).font = _BOLD_FONT
        return start_row + 1

    n_cols = len(df.columns)

    # ── Write all rows ────────────────────────────────────────────────────────
    for r_offset, row_vals in enumerate(
        dataframe_to_rows(df, index=False, header=True)
    ):
        for c_idx, raw_val in enumerate(row_vals, 1):
            val  = _safe_value(raw_val)
            cell = ws.cell(row=start_row + r_offset, column=c_idx, value=val)
            cell.alignment = _LEFT

    # ── Header row styling ────────────────────────────────────────────────────
    _apply_header_row(ws, start_row, n_cols)

    # ── Alternating row fills ─────────────────────────────────────────────────
    data_start = start_row + 1
    data_end   = start_row + len(df)
    _apply_alternating(ws, data_start, data_end, n_cols)

    # ── Column-level formats ──────────────────────────────────────────────────
    for c_idx, col_name in enumerate(df.columns, 1):
        if _is_date_col(col_name, df[col_name].dtype):
            for r in range(data_start, data_end + 1):
                cell = ws.cell(row=r, column=c_idx)
                if cell.value is not None:
                    cell.number_format = _DATE_FMT
        if _is_amount_col(col_name):
            for r in range(data_start, data_end + 1):
                cell = ws.cell(row=r, column=c_idx)
                if cell.value is not None:
                    cell.number_format = _AMOUNT_FMT

    # ── Risk column colour coding ─────────────────────────────────────────────
    if risk_col_name and risk_col_name in df.columns:
        risk_c = list(df.columns).index(risk_col_name) + 1
        for r in range(data_start, data_end + 1):
            cell = ws.cell(row=r, column=risk_c)
            risk_val = str(cell.value).strip() if cell.value else ""
            if risk_val in ("High", "Medium", "Low"):
                cell.fill = _risk_fill(risk_val)
                cell.font = _risk_font_white()

    return data_end + 1


# ---------------------------------------------------------------------------
# Flags column builder (for Full Population sheet)
# ---------------------------------------------------------------------------

def _build_flags_column(
    full_df: pd.DataFrame,
    module_results: dict[str, dict],
) -> pd.Series:
    """
    Build a pipe-separated "Flags" column for the full population sheet.

    For each row in full_df, records which modules flagged it by
    intersecting the module's flagged_df index with the full_df index.

    Args:
        full_df:        Original full-population DataFrame.
        module_results: Dict of module_name → result dict.

    Returns:
        pd.Series indexed like full_df, containing pipe-separated module
        names for flagged rows, empty string for clean rows.
    """
    flags = pd.Series("", index=full_df.index, dtype=str)
    for result in module_results.values():
        flagged_idx = result["flagged_df"].index
        matched     = flagged_idx.intersection(full_df.index)
        name        = result["module_name"]
        for idx in matched:
            existing   = flags[idx]
            flags[idx] = (existing + " | " + name) if existing else name
    return flags


# ---------------------------------------------------------------------------
# Overall risk rating
# ---------------------------------------------------------------------------

def _overall_risk(module_results: dict[str, dict]) -> str:
    """
    Determine the overall engagement risk rating as the worst of all modules.

    Args:
        module_results: Dict of module_name → result dict.

    Returns:
        "High", "Medium", or "Low".
    """
    ratings = [r["risk_rating"] for r in module_results.values()]
    if "High"   in ratings:
        return "High"
    if "Medium" in ratings:
        return "Medium"
    return "Low"


# ---------------------------------------------------------------------------
# Individual sheet writers
# ---------------------------------------------------------------------------

def _write_summary_sheet(
    ws,
    full_df: pd.DataFrame,
    module_results: dict[str, dict],
    params: dict,
) -> None:
    """
    Write Sheet 1 — Summary Dashboard.

    Layout (per CLAUDE.md):
      Row 1:  Workpaper title
      Row 2:  Client / Period / Run Date
      Row 3:  Prepared By / Population / Materiality
      Row 4:  blank
      Row 5:  OVERALL ENGAGEMENT RISK RATING  ← colour-coded
      Row 6:  blank
      Row 7:  Table header: Module | JEs Flagged | % of Population | Risk Rating
      Row 8+: One row per module
      Last:   TOTAL FLAGS row

    Args:
        ws:             Worksheet (already created).
        full_df:        Original full-population DataFrame.
        module_results: Dict of module_name → result dict.
        params:         Engagement params dict (client_name, period,
                        prepared_by, materiality, run_date).
    """
    population  = len(full_df)
    overall     = _overall_risk(module_results)
    run_date    = params.get("run_date", datetime.today())
    date_str    = run_date.strftime(DISPLAY_DATE_FORMAT) if isinstance(run_date, datetime) \
                  else str(run_date)
    materiality = params.get("materiality", "")
    mat_str     = f"₹{materiality:,.0f}" if materiality else "Not specified"

    # ── Row 1: Title ──────────────────────────────────────────────────────────
    ws.cell(row=1, column=1,
            value="KKC & Associates LLP — Journal Entry Testing Workpaper")
    ws.cell(row=1, column=1).font      = _TITLE_FONT
    ws.cell(row=1, column=1).alignment = _LEFT
    ws.column_dimensions["A"].width    = 50

    # ── Row 2: Client / Period / Run Date ─────────────────────────────────────
    ws.cell(row=2, column=1,
            value=f"Client: {params.get('client_name', '')}     "
                  f"Period: {params.get('period', '')}     "
                  f"Run Date: {date_str}")
    ws.cell(row=2, column=1).font      = _BOLD_FONT
    ws.cell(row=2, column=1).alignment = _LEFT

    # ── Row 3: Prepared By / Population / Materiality ─────────────────────────
    ws.cell(row=3, column=1,
            value=f"Prepared By: {params.get('prepared_by', '')}     "
                  f"Total JE Population: {population:,}     "
                  f"Materiality: {mat_str}")
    ws.cell(row=3, column=1).font      = _NORMAL_FONT
    ws.cell(row=3, column=1).alignment = _LEFT

    # ── Row 4: blank ──────────────────────────────────────────────────────────

    # ── Row 5: Overall Risk Rating ────────────────────────────────────────────
    overall_cell = ws.cell(
        row=5, column=1,
        value=f"OVERALL ENGAGEMENT RISK RATING:   {overall.upper()}"
    )
    overall_cell.fill      = _risk_fill(overall)
    overall_cell.font      = Font(name="Arial", size=14, bold=True, color="FFFFFF")
    overall_cell.alignment = _CENTRE
    # Extend colour across columns A–D
    for c in range(2, 5):
        ws.cell(row=5, column=c).fill = _risk_fill(overall)

    # ── Row 6: blank ──────────────────────────────────────────────────────────

    # ── Row 7: Module table header ────────────────────────────────────────────
    headers = ["Module", "JEs Flagged", "% of Population", "Risk Rating"]
    for c_idx, h in enumerate(headers, 1):
        cell           = ws.cell(row=7, column=c_idx, value=h)
        cell.fill      = _GREEN_FILL
        cell.font      = _HEADER_FONT
        cell.alignment = _CENTRE

    # ── Rows 8+: one row per module ───────────────────────────────────────────
    total_flags = 0
    total_pop   = 0
    data_row    = 8

    for module_key, sheet_name in _MODULE_SHEET_MAP:
        result = module_results.get(module_key)
        if result is None:
            continue

        flag_count = result["flag_count"]
        pop_count  = result["population_count"]
        pct        = result["flag_pct"] * 100
        risk       = result["risk_rating"]

        total_flags += flag_count
        total_pop    = max(total_pop, pop_count)  # population is the same across all

        fill = _GREY_FILL if (data_row - 8) % 2 == 1 else _WHITE_FILL

        cells = [
            (1, result["module_name"]),
            (2, flag_count),
            (3, f"{pct:.1f}%"),
            (4, risk),
        ]
        for c_idx, val in cells:
            cell           = ws.cell(row=data_row, column=c_idx, value=val)
            cell.fill      = fill
            cell.font      = _NORMAL_FONT
            cell.alignment = _CENTRE if c_idx > 1 else _LEFT

        # Colour the Risk Rating cell
        risk_cell      = ws.cell(row=data_row, column=4)
        risk_cell.fill = _risk_fill(risk)
        risk_cell.font = _risk_font_white()

        data_row += 1

    # ── Total row ──────────────────────────────────────────────────────────────
    total_pct = (total_flags / total_pop * 100) if total_pop else 0

    for c_idx, val in [
        (1, "TOTAL FLAGS"),
        (2, total_flags),
        (3, f"{total_pct:.1f}%"),
        (4, ""),
    ]:
        cell           = ws.cell(row=data_row, column=c_idx, value=val)
        cell.fill      = _GREEN_FILL
        cell.font      = _HEADER_FONT
        cell.alignment = _CENTRE if c_idx > 1 else _LEFT

    ws.sheet_properties.tabColor = "FFFFFF"
    _freeze_top(ws)
    _autofit(ws)


def _write_benfords_sheet(ws, result: dict) -> None:
    """
    Write Sheet 2 — Benfords Analysis.

    Special layout:
      Section 1: Chi-square test results (key/value block)
      Section 2: Frequency table (digit-level expected vs observed)
      Section 3: Flagged individual entries

    Args:
        ws:     Worksheet.
        result: Benfords module result dict.
    """
    stats   = result.get("summary_stats", {})
    chi2    = stats.get("chi2", "N/A")
    p_value = stats.get("p_value", "N/A")
    dof     = stats.get("degrees_of_freedom", "N/A")
    pf      = stats.get("pass_fail", "N/A")
    risk    = result.get("risk_rating", "Low")

    # ── Section 1: Chi-square stats ───────────────────────────────────────────
    _section_header(ws, 1, "Chi-Square Goodness-of-Fit Test Results", n_cols=4)
    kv_pairs = [
        ("Chi-Square Statistic",  round(chi2,    4) if isinstance(chi2,    float) else chi2),
        ("P-Value",               round(p_value, 6) if isinstance(p_value, float) else p_value),
        ("Degrees of Freedom",    dof),
        ("Test Result",           pf),
        ("Risk Rating",           risk),
        ("Population (analysed)", result.get("population_count", 0)),
        ("Entries Flagged",       result.get("flag_count",       0)),
    ]
    row_after_kv = _write_kv_block(ws, kv_pairs, 2)

    # ── Section 2: Frequency table ────────────────────────────────────────────
    freq_table: pd.DataFrame = stats.get("freq_table", pd.DataFrame())
    row_after_kv += 1   # blank row

    _section_header(ws, row_after_kv, "Benford's Law Frequency Table", n_cols=7)
    row_after_kv += 1

    if not freq_table.empty:
        display_ft = freq_table.copy()
        display_ft["expected_pct"] = (display_ft["expected_pct"] * 100).round(2)
        display_ft["observed_pct"] = (display_ft["observed_pct"] * 100).round(2)
        display_ft["deviation_pct"] = display_ft["deviation_pct"].round(2)
        display_ft.columns = [
            "Digit", "Expected %", "Expected Count",
            "Observed Count", "Observed %", "Deviation", "Deviation %",
        ]
        row_after_freq = _write_df(ws, display_ft, row_after_kv)
    else:
        ws.cell(row=row_after_kv, column=1, value="(Frequency table unavailable)")
        row_after_freq = row_after_kv + 1

    # ── Section 3: Flagged entries ────────────────────────────────────────────
    row_after_freq += 1
    _section_header(ws, row_after_freq, "Flagged Entries (Over-Represented Digit Buckets)",
                    n_cols=max(len(result["flagged_df"].columns), 1))
    _write_df(ws, result["flagged_df"], row_after_freq + 1)

    ws.sheet_properties.tabColor = _tab_colour(risk)
    _freeze_top(ws)
    _autofit(ws)


def _write_standard_module_sheet(ws, result: dict) -> None:
    """
    Write a standard module sheet (Sheets 3–8, 10):
      - Small stats block at top
      - Full flagged_df table below

    Args:
        ws:     Worksheet.
        result: Standard module result dict.
    """
    flag_count = result["flag_count"]
    pop_count  = result["population_count"]
    pct        = result["flag_pct"] * 100
    risk       = result["risk_rating"]
    mod_name   = result["module_name"]

    # ── Stats block ───────────────────────────────────────────────────────────
    kv_pairs = [
        ("Module",            mod_name),
        ("Population",        pop_count),
        ("Entries Flagged",   flag_count),
        ("% of Population",   f"{pct:.2f}%"),
        ("Risk Rating",       risk),
    ]
    row_after = _write_kv_block(ws, kv_pairs, 1)

    # Colour risk rating cell in the stats block (row 5, col 2)
    risk_cell       = ws.cell(row=5, column=2)
    risk_cell.fill  = _risk_fill(risk)
    risk_cell.font  = _risk_font_white()

    # ── Flagged entries table ─────────────────────────────────────────────────
    row_after += 1
    n_flag_cols = max(len(result["flagged_df"].columns), 1)
    _section_header(ws, row_after, "Flagged Entries", n_cols=n_flag_cols)
    _write_df(ws, result["flagged_df"], row_after + 1)

    ws.sheet_properties.tabColor = _tab_colour(risk)
    _freeze_top(ws)
    _autofit(ws)


def _write_user_analysis_sheet(ws, result: dict) -> None:
    """
    Write Sheet 9 — User Analysis.

    Special layout with three sections:
      Section 1: Stats block (risk, counts, outlier thresholds)
      Section 2: Top-10 users by entry volume
      Section 3: Top-10 users by total amount
      Section 4: Full flagged entries (self-approval, outliers, etc.)

    Args:
        ws:     Worksheet.
        result: User Analysis module result dict.
    """
    stats  = result.get("summary_stats", {})
    risk   = result.get("risk_rating", "Low")
    pop    = result.get("population_count", 0)

    # ── Section 1: Stats ──────────────────────────────────────────────────────
    vol_thresh = stats.get("volume_outlier_threshold", 0)
    amt_thresh = stats.get("amount_outlier_threshold", 0)
    vol_thresh_str = f"{vol_thresh:.1f}" if isinstance(vol_thresh, float) else str(vol_thresh)
    amt_thresh_str = (f"₹{amt_thresh:,.0f}"
                      if isinstance(amt_thresh, (int, float)) and amt_thresh < 1e15
                      else "N/A")

    kv_pairs = [
        ("Module",                    "User Analysis"),
        ("Population",                pop),
        ("Entries Flagged",           result.get("flag_count", 0)),
        ("% of Population",           f"{result.get('flag_pct', 0)*100:.2f}%"),
        ("Risk Rating",               risk),
        ("Self-Approvals",            stats.get("self_approval_count", 0)),
        ("Volume Outlier Users",      ", ".join(stats.get("volume_outlier_users", []))),
        ("Amount Outlier Users",      ", ".join(stats.get("amount_outlier_users", []))),
        ("Single-User Day Entries",   stats.get("single_user_day_count", 0)),
        ("Volume Outlier Threshold",  vol_thresh_str + " entries"),
        ("Amount Outlier Threshold",  amt_thresh_str),
    ]
    row_after = _write_kv_block(ws, kv_pairs, 1)

    risk_cell      = ws.cell(row=5, column=2)
    risk_cell.fill = _risk_fill(risk)
    risk_cell.font = _risk_font_white()

    # ── Section 2: Top-10 by volume ───────────────────────────────────────────
    row_after += 1
    top10_vol: pd.DataFrame = stats.get("top10_by_volume", pd.DataFrame())
    n_vol_cols = max(len(top10_vol.columns) if not top10_vol.empty else 1, 1)
    _section_header(ws, row_after, "Top 10 Users by Entry Volume", n_cols=n_vol_cols)
    row_after = _write_df(ws, top10_vol, row_after + 1) + 1

    # ── Section 3: Top-10 by amount ───────────────────────────────────────────
    top10_amt: pd.DataFrame = stats.get("top10_by_amount", pd.DataFrame())
    n_amt_cols = max(len(top10_amt.columns) if not top10_amt.empty else 1, 1)
    _section_header(ws, row_after, "Top 10 Users by Total Amount Posted", n_cols=n_amt_cols)
    row_after = _write_df(ws, top10_amt, row_after + 1) + 1

    # ── Section 4: Flagged entries ────────────────────────────────────────────
    n_flag_cols = max(len(result["flagged_df"].columns) if not result["flagged_df"].empty else 1, 1)
    _section_header(ws, row_after, "Flagged Entries", n_cols=n_flag_cols)
    _write_df(ws, result["flagged_df"], row_after + 1)

    ws.sheet_properties.tabColor = _tab_colour(risk)
    _freeze_top(ws)
    _autofit(ws)


def _write_full_population_sheet(
    ws,
    full_df: pd.DataFrame,
    module_results: dict[str, dict],
) -> None:
    """
    Write Sheet 11 — Full Population.

    Writes the complete original dataset with an additional "Flags" column
    showing which modules flagged each entry (pipe-separated).

    Args:
        ws:             Worksheet.
        full_df:        Original full-population DataFrame.
        module_results: Dict of module_name → result dict.
    """
    flags_col = _build_flags_column(full_df, module_results)
    display   = full_df.copy()
    display["Flags"] = flags_col

    _write_df(ws, display, start_row=1)

    ws.sheet_properties.tabColor = "FFFFFF"
    _freeze_top(ws)
    _autofit(ws)


def _write_parameters_sheet(
    ws,
    params: dict,
    module_results: dict[str, dict],
) -> None:
    """
    Write Sheet 12 — Parameters.

    Records all run parameters and thresholds used in the analysis,
    providing an audit trail per ICAI / NFRA requirements.

    Args:
        ws:             Worksheet.
        params:         Engagement params dict.
        module_results: Dict of module_name → result dict (for module counts).
    """
    run_date = params.get("run_date", datetime.today())
    date_str = (run_date.strftime(DISPLAY_DATE_FORMAT)
                if isinstance(run_date, datetime) else str(run_date))

    try:
        from config import (
            PERIOD_END_HIGH_DAYS, PERIOD_END_MEDIUM_DAYS,
            BUSINESS_HOURS_START, BUSINESS_HOURS_END,
            ROUND_MINIMUM_AMOUNT, RARE_COMBO_THRESHOLD,
            USER_OUTLIER_STD, BENFORD_SIGNIFICANCE,
        )
    except ImportError:
        PERIOD_END_HIGH_DAYS   = 3
        PERIOD_END_MEDIUM_DAYS = 5
        BUSINESS_HOURS_START   = 9
        BUSINESS_HOURS_END     = 19
        ROUND_MINIMUM_AMOUNT   = 10000
        RARE_COMBO_THRESHOLD   = 3
        USER_OUTLIER_STD       = 2.0
        BENFORD_SIGNIFICANCE   = 0.05

    overall = _overall_risk(module_results)
    total_flags = sum(r["flag_count"] for r in module_results.values())
    pop = next(iter(module_results.values()))["population_count"] if module_results else 0

    kv_pairs = [
        ("ENGAGEMENT DETAILS",          ""),
        ("Client Name",                 params.get("client_name", "")),
        ("Audit Period",                params.get("period", "")),
        ("Financial Year",              params.get("financial_year", "")),
        ("Prepared By",                 params.get("prepared_by", "")),
        ("Materiality (₹)",             params.get("materiality", "Not specified")),
        ("Run Date",                    date_str),
        ("Tool Version",                "1.0.0"),
        ("",                            ""),
        ("ANALYSIS SUMMARY",            ""),
        ("Total JE Population",         pop),
        ("Total Entries Flagged",       total_flags),
        ("% Flagged",                   f"{total_flags/pop*100:.2f}%" if pop else "0.00%"),
        ("Overall Risk Rating",         overall),
        ("",                            ""),
        ("THRESHOLDS APPLIED",          ""),
        ("Benford's p-value threshold", BENFORD_SIGNIFICANCE),
        ("Period-End High Band (days)",  PERIOD_END_HIGH_DAYS),
        ("Period-End Medium Band (days)", PERIOD_END_MEDIUM_DAYS),
        ("Business Hours Start",        f"{BUSINESS_HOURS_START:02d}:00"),
        ("Business Hours End",          f"{BUSINESS_HOURS_END:02d}:00"),
        ("Round Number Minimum (₹)",    ROUND_MINIMUM_AMOUNT),
        ("Rare Combo Threshold",        f"< {RARE_COMBO_THRESHOLD} occurrences"),
        ("User Outlier Std Deviations", USER_OUTLIER_STD),
    ]

    for r_idx, (label, value) in enumerate(kv_pairs, 1):
        lc = ws.cell(row=r_idx, column=1, value=label)
        vc = ws.cell(row=r_idx, column=2, value=value)
        lc.alignment = _LEFT
        vc.alignment = _LEFT
        # Section headers (all-caps labels with empty value)
        if label in ("ENGAGEMENT DETAILS", "ANALYSIS SUMMARY", "THRESHOLDS APPLIED"):
            lc.fill = _GREEN_FILL
            lc.font = _HEADER_FONT
            vc.fill = _GREEN_FILL
            vc.font = _HEADER_FONT
            for c in range(3, 7):
                ws.cell(row=r_idx, column=c).fill = _GREEN_FILL
        elif label:
            lc.font = _BOLD_FONT
            vc.font = _NORMAL_FONT

    # Colour the overall risk cell
    for r_idx, (label, _) in enumerate(kv_pairs, 1):
        if label == "Overall Risk Rating":
            cell       = ws.cell(row=r_idx, column=2)
            cell.fill  = _risk_fill(overall)
            cell.font  = _risk_font_white()

    ws.sheet_properties.tabColor = "FFFFFF"
    _freeze_top(ws)
    _autofit(ws)


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------

def generate_filename(
    client_name: str,
    period: str,
    run_date: datetime | None = None,
) -> str:
    """
    Generate the standard KKC workbook filename.

    Format: KKC_JE_Testing_[ClientName]_[Period]_[DDMMYYYY].xlsx
    Example: KKC_JE_Testing_ShriramFinance_Mar2025_15042025.xlsx

    Spaces in client_name and period are replaced with underscores.
    Special characters are stripped.

    Args:
        client_name: Client/company name.
        period:      Audit period string, e.g. "March 2025".
        run_date:    Date to embed in filename; defaults to today.

    Returns:
        Filename string (not a full path).
    """
    if run_date is None:
        run_date = datetime.today()

    def _clean(s: str) -> str:
        import re
        return re.sub(r"[^A-Za-z0-9]", "", s.replace(" ", ""))

    date_part   = run_date.strftime("%d%m%Y")
    client_part = _clean(client_name)
    period_part = _clean(period)
    return f"{OUTPUT_PREFIX}Testing_{client_part}_{period_part}_{date_part}.xlsx"


def build_workbook(
    full_df: pd.DataFrame,
    module_results: dict[str, dict],
    params: dict,
    output_path,
) -> str:
    """
    Build and save the 12-sheet KKC JE Testing workpaper.

    Each module sheet receives the flagged entries from the corresponding
    module result dict.  Sheet tabs are colour-coded by risk rating.
    The full population sheet includes a pipe-separated Flags column.

    Args:
        full_df:        Original full-population DataFrame (all rows as
                        uploaded, after column mapping).
        module_results: Dict of module_name → standard result dict.
                        Keys must match the module_name values in
                        _MODULE_SHEET_MAP.  Missing modules are skipped
                        gracefully.
        params:         Engagement parameter dict with keys:
                          client_name  (str)
                          period       (str)
                          financial_year (str, optional)
                          prepared_by  (str)
                          materiality  (float, optional)
                          run_date     (datetime, optional — defaults to today)
        output_path:    Full file path to save the workbook to
                        (including filename and .xlsx extension).

    Returns:
        The output_path string (allows caller to display/open the file).

    Raises:
        OSError: If the output directory does not exist or is not writable.
    """
    wb = Workbook()
    # Remove the default empty sheet
    wb.remove(wb.active)

    # ── Sheet 1: Summary Dashboard ────────────────────────────────────────────
    ws_summary = wb.create_sheet("Summary Dashboard")
    _write_summary_sheet(ws_summary, full_df, module_results, params)

    # ── Sheets 2–10: Module sheets ────────────────────────────────────────────
    for module_key, sheet_name in _MODULE_SHEET_MAP:
        result = module_results.get(module_key)
        ws     = wb.create_sheet(sheet_name)

        if result is None:
            ws.cell(row=1, column=1, value=f"Module '{module_key}' was not run.")
            ws.cell(row=1, column=1).font = _BOLD_FONT
            ws.sheet_properties.tabColor = "FFFFFF"
            continue

        if module_key == "Benford's Law":
            _write_benfords_sheet(ws, result)
        elif module_key == "User Analysis":
            _write_user_analysis_sheet(ws, result)
        else:
            _write_standard_module_sheet(ws, result)

    # ── Sheet 11: Full Population ─────────────────────────────────────────────
    ws_pop = wb.create_sheet("Full Population")
    _write_full_population_sheet(ws_pop, full_df, module_results)

    # ── Sheet 12: Parameters ──────────────────────────────────────────────────
    ws_params = wb.create_sheet("Parameters")
    _write_parameters_sheet(ws_params, params, module_results)

    # ── Save ──────────────────────────────────────────────────────────────────
    wb.save(output_path)
    return output_path
